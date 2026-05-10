import time
import os
import sys
import cv2
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
#from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage


# ==============================
# Fix path for EXE
# ==============================
def resource_path(file_name):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, file_name)
    return os.path.join(os.path.abspath("."), file_name)


# ==============================
# Create Driver
# ==============================
def create_driver(update_status):
    try:
        driver_path = resource_path("chromedriver.exe")

        if not os.path.exists(driver_path):
            raise Exception("chromedriver.exe not found")

        service = Service(driver_path)

        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")

        driver = webdriver.Chrome(service=service, options=options)
        return driver

    except Exception as e:
        update_status(f"Driver Error: {str(e)}")
        raise


# ==============================
# MAIN AUTOMATION
# ==============================
def run_automation(output_folder, excel_path, update_status, ask_login):
    try:
        # Clean path
        output_folder = output_folder.strip()

        # Ensure output folder exists
        os.makedirs(output_folder, exist_ok=True)

        # Output Excel path (DO NOT overwrite input file!)
        output_excel_path = os.path.join(output_folder, "output.xlsx")

        update_status("🚀 Starting browser...")

        driver = create_driver(update_status)
        wait = WebDriverWait(driver, 40)

        base_url = "https://haside9.ext.hitachiastemo.com/qm/web/console/PJ_INV_VW_MEBperf_SWE/_WwcSEG3fEfCYdpgxNQni9A#action=com.ibm.rqm.planning.home.actionDispatcher&subAction=viewTestCases&updateAction=clear-filter-selection"

        driver.get(base_url)

        # Manual login
        ask_login()

        update_status("📄 Reading ETM IDs...")

        # ==============================
        # READ INPUT EXCEL (USER FILE)
        # ==============================
        wb_input = load_workbook(excel_path)
        ws_input = wb_input.active

        etm_ids = [
            str(row[0])
            for row in ws_input.iter_rows(values_only=True)
            if row[0]
        ]

        # ==============================
        # PREPARE OUTPUT EXCEL
        # ==============================
        if os.path.exists(output_excel_path):
            wb = load_workbook(output_excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Traceability"
            ws["A1"] = "#7"

        total = len(etm_ids)

        # ==============================
        # LOOP
        # ==============================
        for i, etm_id in enumerate(etm_ids, start=1):
            update_status(f"🔄 Processing {i}/{total}: {etm_id}")

            try:
                driver.get(base_url)
                driver.refresh()
                time.sleep(3)

                # Filter
                filter_box = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//input[@title='Type filter text and press Enter']")
                    )
                )

                filter_box.clear()
                filter_box.send_keys(etm_id)
                filter_box.send_keys(Keys.ENTER)
                time.sleep(3)

                # Click ETM ID
                id_element = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, f"//a[contains(@href,'id={etm_id}')]")
                    )
                )

                driver.execute_script("arguments[0].click();", id_element)
                time.sleep(3)

                # Wait page
                wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//ul[contains(@class,'entries')]")
                    )
                )

                # Click Execution tab
                exec_tab = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//ul[@class='entries']//li[contains(@id,'execution')]")
                    )
                )
                exec_tab.click()
                time.sleep(5)

                # Screenshot 1
                driver.execute_script("document.body.style.zoom='80%'")
                driver.execute_script("window.scrollTo(0, 60);")

                img1_path = os.path.join(output_folder, f"execution_{etm_id}.png")
                driver.save_screenshot(img1_path)

                img = cv2.imread(img1_path)
                x1 = 260
                y1 = 628

                x2 = 1590
                y2 = 700
                cv2.rectangle(
            img,
            (x1, y1),
            (x2, y2),
            (0, 0, 255),
            4
        )

                cv2.imwrite(img1_path, img)
                update_status("✅ Previous Result Highlight Added")



                # Click Name
                elements = wait.until(
                    EC.presence_of_all_elements_located(
                        (By.XPATH, "//table[contains(@class,'content-table')]//a[contains(@class,'jazz-ui-ResourceLink')]")
                    )
                )

                visible_links = [el for el in elements if el.is_displayed()]

                if len(visible_links) < 6:
                    raise Exception("Not enough visible links found")

                name_element = visible_links[5]
                driver.execute_script("arguments[0].click();", name_element)
                time.sleep(5)

                # Previous Result
                prev_result = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//a[@title='Previous Result Details']")
                    )
                )

                driver.execute_script("arguments[0].click();", prev_result)
                time.sleep(3)

                driver.execute_script("window.scrollTo(0, 60);")

                img2_path = os.path.join(output_folder, f"previous_{etm_id}.png")
                driver.save_screenshot(img2_path)
 
                img = cv2.imread(img2_path)
                       # CHANGE THESE COORDINATES
                x1 = 260
                y1 = 628
                x2 = 1590
                y2 = 700
                cv2.rectangle(
                    img,
                    (x1, y1),
                    (x2, y2),
                    (0, 0, 255),
                    4
                    )
                cv2.imwrite(img2_path, img)
                update_status("✅ Previous Result Highlight Added")

        

                # ==============================
                # SAVE TO EXCEL
                # ==============================
                row = ws.max_row + 2
                ws[f"A{row}"] = " "
                ws.row_dimensions[row].height = 200

                img1 = XLImage(img1_path)
                img1.width = 600
                img1.height = 250
                ws.add_image(img1, f"B{row}")

                img2 = XLImage(img2_path)
                img2.width = 600
                img2.height = 250
                ws.add_image(img2, f"D{row}")

                ws.column_dimensions['B'].width = 90
                ws.column_dimensions['D'].width = 90

                wb.save(output_excel_path)

                update_status(f"✅ Done {etm_id}")

            except Exception as e:
                update_status(f"❌ Error in {etm_id}: {str(e)}")

        driver.quit()
        update_status("🎉 Automation Completed!")

    except Exception as e:
        update_status(f"🔥 Fatal Error: {str(e)}")